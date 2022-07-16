import openpyxl
from pathlib import Path


class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        """
        初始化函数，解析Excel文件获取Sheet以及Sheet中最大行数
        :param excelPath:
        :param sheetName:
        """
        self.wb = openpyxl.load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        """
        遍历每行数据放到List，并返回List， 之后可以通过索引获取
        :return:
        """
        dataList = []
        for line in list(self.sheet.rows)[1:]:
            tmpList = [line[1].value]
            # tmpList.append(line[2].value)
            dataList.append(tmpList)
        return dataList


class ReadExcel:

    def __init__(self, file):
        self.file = file
        self.fb = openpyxl.load_workbook(self.file)
        self.sheet = self.fb.active

    # 获取所有行
    def get_rows(self):
        num = 1
        while True:
            cell = self.sheet.cell(row=num, column=1).value
            if cell:
                num = num + 1
            else:
                return num - 1

    # 获取某个单元格数据
    def get_value(self, row, col):
        value = self.sheet.cell(row, col).value
        return value

    # 在某一个单元格写入数据
    def write_value(self, row, col, value):
        self.sheet.cell(row=row, column=col).value = value
        self.fb.save(self.file)

